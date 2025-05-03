import json
import openpyxl
from openpyxl.styles import Border, Side, Alignment

# Открываем файл с данными
with open('./busyRoomsResult4.txt', 'r', encoding='utf-8') as file:
    data = json.load(file)

# Фильтруем данные по weekType
filtered_data = [item for item in data if item.get('weekType') is not None]

# Создаем словарь для хранения данных по дням недели и времени
schedule = {}

# Проходим по каждому элементу и группируем по дням недели
for item in filtered_data:
    day_of_week = item.get('dayOfWeek')
    time_slot_id = item.get('timeSlotId')
    time_slot_title = item.get('timeSlot', {}).get('title')

    if day_of_week not in schedule:
        schedule[day_of_week] = {}

    if time_slot_id not in schedule[day_of_week]:
        schedule[day_of_week][time_slot_id] = time_slot_title

# Создаем новый файл Excel
wb = openpyxl.Workbook()
ws = wb.active

# Определяем границы
thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'),
                     bottom=Side(style='thin'))

# Очень тонкие границы для дисциплин
very_thin_border = Border(left=Side(style='dotted'), right=Side(style='dotted'), top=Side(style='dotted'),
                          bottom=Side(style='dotted'))

# Вставляем заголовки в ячейки A2 и B2 (сдвигаем их на одну строку вниз)
ws.cell(row=3, column=1, value="Дни")
ws.cell(row=3, column=2, value="Часы")

# Применяем тонкую границу для заголовков в ячейках A2 и B2
ws.cell(row=3, column=1).border = thin_border
ws.cell(row=3, column=2).border = thin_border

# Извлекаем уникальные группы и их направления
unique_groups = set()
group_directions = {}  # Для хранения направлений по группам
group_titles = {}
for item in filtered_data:
    for group in item.get('groups', []):
        profile = group['group']['profile']
        title = group['group']['title']
        direction = group['group']['direction']
        unique_groups.add(profile)  # Заменено на profile
        if direction not in group_directions:
            group_directions[direction] = set()  # Используем set для уникальных профилей
            group_titles[direction] = set()
        group_directions[direction].add(profile)  # Добавляем уникальные профили в соответствующее направление
        group_titles[direction].add(title)

# Записываем уникальные группы в Excel начиная с D2
group_column = 4  # Начинаем с D2 (столбец 4)
for direction, profiles in group_directions.items():
    # Объединяем ячейки для direction и записываем значение в строку 1
    end_column = group_column + len(profiles) - 1
    ws.merge_cells(start_row=1, start_column=group_column, end_row=1, end_column=end_column)
    ws.cell(row=1, column=group_column, value=direction)
    ws.cell(row=1, column=group_column).border = thin_border

    # Записываем уникальные профили в соответствующие ячейки
    for i, profile in enumerate(profiles):
        ws.cell(row=2, column=group_column + i, value=profile)
        ws.cell(row=2, column=group_column + i).border = thin_border

    titles = list(group_titles[direction])
    for i, title in enumerate(titles):
        ws.cell(row=3, column=group_column + i, value=title)
        ws.cell(row=3, column=group_column + i).border = thin_border

    group_column += len(profiles)  # Переходим к следующей группе

# Полные названия дней недели
days_of_week_full = {
    "MON": "ПОНЕДЕЛЬНИК",
    "TUE": "ВТОРНИК",
    "WED": "СРЕДА",
    "THU": "ЧЕТВЕРГ",
    "FRI": "ПЯТНИЦА",
    "SAT": "СУББОТА"
}

# Заполняем таблицу временными интервалами и объединяем ячейки
row = 4  # Начинаем с 3-й строки, так как строка 1 пустая, а строка 2 - заголовки
for day, time_slots in schedule.items():  # Перебираем каждый день недели
    start_row = row
    end_row = row + len(time_slots) * 2 - 1  # Объединяем все строки для дня недели в столбце A
    ws.merge_cells(start_row=start_row, start_column=1, end_row=end_row,
                   end_column=1)  # Объединяем ячейки в столбце A для дня недели
    # Вставляем название дня недели в объединенную ячейку
    full_day_name = days_of_week_full.get(day, day)
    ws.cell(row=start_row, column=1, value=full_day_name)
    # Применяем жирную границу для дня недели в столбце A
    for i in range(start_row, end_row + 1):
        ws.cell(row=i, column=1).border = thin_border

    # Поворачиваем текст на 90 градусов
    ws.cell(row=start_row, column=1).alignment = Alignment(textRotation=90, horizontal='center', vertical='center')

    # Заполняем временные слоты в столбце B и добавляем четную и нечетную неделю в столбец C
    for time_slot, time_title in time_slots.items():
        ws.merge_cells(start_row=row, start_column=2, end_row=row + 1,
                       end_column=2)  # Объединяем ячейки для времени в столбце B
        ws.cell(row=row, column=2, value=time_title)
        # Применяем тонкую границу для ячеек с временем в столбце B
        for i in range(row, row + 2):
            ws.cell(row=i, column=2).border = thin_border
        # Выравнивание для времени
        for i in range(row, row + 2):
            ws.cell(row=i, column=2).alignment = Alignment(horizontal='center', vertical='center')

        # Вставляем "Нечетная неделя" и "Четная неделя" в столбец C
        ws.cell(row=row, column=3, value="Нечетная неделя")
        ws.cell(row=row + 1, column=3, value="Четная неделя")
        # Применяем тонкую границу для ячеек с неделями в столбце C
        for i in range(row, row + 2):
            ws.cell(row=i, column=3).border = thin_border
        # Выравнивание для недели
        for i in range(row, row + 2):
            ws.cell(row=i, column=3).alignment = Alignment(horizontal='center', vertical='center')

        # Заполняем дисциплины в соответствующие ячейки
        for item in filtered_data:
            for group in item.get('groups', []):
                if group['group']['profile'] in unique_groups:  # Используем profile вместо title
                    if item['dayOfWeek'] == day and item['timeSlotId'] == time_slot:
                        week_type = item['weekType']
                        column_idx = list(unique_groups).index(group['group']['profile']) + 4

                        # Форматируем дисциплину
                        discipline_type = "лек." if item['assignment']['type'] == "lecture" else "пр."
                        discipline_name = item['assignment']['discipline']

                        # Проверяем наличие преподавателя и добавляем department и position
                        teacher_name = "Не указан"
                        department_name = "Не указан"
                        position_name = "Не указан"
                        if item['teachers']:
                            teacher = item['teachers'][0]['teacher']['user']
                            department_name = item['teachers'][0]['teacher']['department']['title']
                            position_name = item['teachers'][0]['teacher']['position']['title']
                            teacher_name = f"{teacher['lastName']} {teacher['firstName'][0]}. {teacher['middleName'][0] if teacher.get('middleName') else ''}."

                        # Получаем название аудитории
                        audience = item['rooms'][0]['audience']['title'] if item['rooms'] else "Не указано"

                        # Формируем строку
                        full_entry = f"{discipline_type} {discipline_name}\n{teacher_name}, {department_name}, {position_name}\n{audience}"

                        # Записываем дисциплину с более тонкими границами для дисциплин
                        if week_type == 'EVEN':
                            ws.cell(row=row, column=column_idx, value=full_entry)
                            ws.cell(row=row, column=column_idx).border = thin_border
                        elif week_type == 'ODD':
                            ws.cell(row=row + 1, column=column_idx, value=full_entry)
                            ws.cell(row=row + 1, column=column_idx).border = thin_border

        row += 2  # Переходим ко второму интервалу

# Автоматическая ширина столбцов
for col in ws.columns:
    max_length = 0
    column = col[0]  # Получаем первую ячейку столбца
    if not isinstance(column, openpyxl.cell.MergedCell):  # Проверяем, не объединенная ли ячейка
        column_letter = column.column_letter  # Получаем букву столбца
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2)  # Добавляем немного пространства
        ws.column_dimensions[column_letter].width = adjusted_width

# Сохраняем файл
wb.save('output_schedule.xlsx')

print("Файл успешно создан и сохранен как 'output_schedule.xlsx'")
